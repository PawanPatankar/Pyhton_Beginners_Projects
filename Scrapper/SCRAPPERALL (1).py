import requests
from bs4 import BeautifulSoup
import threading
import  csv
import platform, time, urllib.request
from selenium import webdriver
import os, time, openpyxl, operator, tweepy
from openpyxl import Workbook


def key1fil1(txt1,oper,date2,date3,limit):
    response = requests.get(f"https://api.data.io/v2/search?query=language%3Aen%20AND%20text%3A%22{txt1[0]}%22%20{oper}%20text%3A%22{txt1[2]}%22%20AND%20discoverDate%3A%5B{date2}%20TO%20{date3}%5D&sortBy=discoverDate&sortOrder=DESC&limit={limit}",headers={"Authorization":"sBBqsGXiYgF0Db5OV5tAw6SAtFG7TDb-b4qTEp89VCPcWZ-teGVG-eZRAeTThSiLn2pHZrSf1gT2PUujH1YaQA"})
    jsonf = response.json()
    return jsonf

def key1fill2(txt1,oper,limit,):
    response = requests.get(f"https://api.data.io/v2/search?query=language%3Aen%20AND%20text%3A%22{txt1[0]}%22%20{oper}%20text%3A%22{txt1[2]}%22&sortBy=_score&sortOrder=DESC&limit={limit}",headers={"Authorization":"sBBqsGXiYgF0Db5OV5tAw6SAtFG7TDb-b4qTEp89VCPcWZ-teGVG-eZRAeTThSiLn2pHZrSf1gT2PUujH1YaQA"})
    jsonf = response.json()
    return jsonf

def key2fill1(text,date2,date3,limit):
    txtdat = requests.get(f"https://api.data.io/v2/search?query=language%3Aen%20AND%20text%3A{text}%20AND%20discoverDate%3A%5B{date2}%20TO%20{date3}%5D&sortBy=_score&sortOrder=DESC&limit={limit}",headers={"Authorization": "sBBqsGXiYgF0Db5OV5tAw6SAtFG7TDb-b4qTEp89VCPcWZ-teGVG-eZRAeTThSiLn2pHZrSf1gT2PUujH1YaQA"})
    jsonf = txtdat.json()
    return jsonf

def key2fill2(text,limit):
    ontxt = requests.get(f"https://api.data.io/v2/search?query=language%3Aen%20AND%20text%3A{text}&sortBy=_score&sortOrder=DESC&limit={limit}", headers={"Authorization": "sBBqsGXiYgF0Db5OV5tAw6SAtFG7TDb-b4qTEp89VCPcWZ-teGVG-eZRAeTThSiLn2pHZrSf1gT2PUujH1YaQA"})
    jsonf = ontxt.json()
    return jsonf




# ---------------------------------------------------------------------------------------------------------------------------------------------------

def keygdateg1(keygi):
    responseg = requests.get(f"https://news.google.com/search?q={keygi}%20when%3A1h&hl=en-IN&gl=IN&ceid=IN%3Aen").text
    return responseg

def keygdateg2(keygi):
    responseg = requests.get(f"https://news.google.com/search?q={keygi}%20when%3A1d&hl=en-IN&gl=IN&ceid=IN%3Aen").text
    return responseg

def keygdateg3(keygi):
    responseg = requests.get(f"https://news.google.com/search?q={keygi}%20when%3A7d&hl=en-IN&gl=IN&ceid=IN%3Aen").text
    return responseg
def keygdateg4(keygi):
    responseg = requests.get(f"https://news.google.com/search?q={keygi}%20when%3A1y&hl=en-IN&gl=IN&ceid=IN%3Aen").text
    return responseg
def keyg2r(keygi):
    # responseg = requests.get(f"https://news.google.com/search?q={keygi}&hl=en-IN&gl=IN&ceid=IN%3Aen").text
    responseg = requests.get(f"https://news.google.com/search?q={keygi}&hl=en-IN&gl=IN&ceid=IN%3Aen").text
    return responseg





def forg(soup):
    csv_file2 = open(f'{keygi}.csv', 'a',encoding='utf-8')

    csv_writer = csv.writer(csv_file2)
    csv_writer.writerow(['Headline', 'Link'])
    artic=[]
    lin = []
    file2 = open(f'{keygi}.txt', 'a',encoding='utf-8')
    for link in soup.find_all('a',class_='DY5T1d', href=True):
        # print ("https://news.google.com/"+link['href'])
        artic.append(link.text)
        lin.append("https://news.google.com/"+link['href'])

    for i,j in zip(artic,lin):

        print(i)
        file2.write(f"Headline:\n{i}\n\n")
        print(j)
        file2.write(f"Link:\n{j}\n\n")
        print("-------------------------   GOOGLE  --------------------------------")
        file2.write("------------------------------  GOOGLE  --------------------------------------")
        csv_writer.writerow([i, j])
#MAinFILE
#MAinFILE
def forn(jsonf):

    csv_file = open(f'{text1}.csv', 'a',encoding='utf-8')

    csv_writer = csv.writer(csv_file)
    csv_writer.writerow(['Title', 'Date', 'Article'])
    file = open(f'{text1}.txt', 'a',encoding='utf-8')
    for item in jsonf:

        title = item.get('title')
        file.write(f"{title}\n \n")
        print(title)
        date = item.get('discoverDate')
        file.write(f"Date: {date}\n\n")
        print(date)
        text = item.get('text')
        file.write(f"Article \n\n {text} \n \n")
        print(text)
        print("-----------------------------------  NEWS RIVER  ----------------------------------------------------\n")
        file.write("-----------------------------------------  NEWS RIVER  --------------------------------------------------\n")
        csv_writer.writerow([title, date, text])

def Create_Dir(dir_name):
        if not os.path.exists("data"):
            try:
                os.mkdir("data")
                print("Created directory 'data'")
            except:
                print("Unable to create directory 'data': Directory already exists")
        else:
            print("Unable to create directory 'data': Directory already exists")

        if not os.path.exists("data/data_" + dir_name):
            try:
                os.mkdir("data/data_" + dir_name)
                print("Created directory 'data/data_" + dir_name + "'")
            except:
                print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")
        else:
            print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")

        if not os.path.exists("data/data_" + dir_name + '/img'):
            try:
                os.mkdir("data/data_" + dir_name + '/img')
                print("Created directory 'data/data_" + dir_name + "/img'")
            except:
                print("Unable to create directory 'data/data_" + dir_name + "/img': Directory already exists")
        else:
            print("Unable to create directory 'data/data_" + dir_name + "/img': Directory already exists")

def Scrape_Instagram(tag, limit, browser):
        Create_Dir(tag)

        print("Starting Scrapping Instagram")
        file_path = "data/data_" + tag
        keyword = tag
        # Adding path.
        if not os.getcwd() in os.get_exec_path():
            # print('adding path')
            if platform.system() == "Windows":
                os.environ["PATH"] = os.environ["PATH"] + ";" + os.getcwd()
            else:
                os.environ["PATH"] = os.environ["PATH"] + ":" + os.getcwd()

        # opening instagram in browser
        if 'chrome' in browser.lower():
            driver = webdriver.Chrome()
        else:
            driver = webdriver.Firefox()
        driver.get("https://www.instagram.com/" + "explore/tags/" + tag)

        print("Loading Posts")
        time.sleep(10)
        print("Loading Data")

        time.sleep(5)

        print(str(limit) + " images loaded")


        img_src = []

        for data in driver.find_elements_by_class_name("FFVAD"):

            img_src.append(data.get_attribute("src"))

        img_src = img_src[9:limit + 9]

        print("Dumping Images. This will take some time!")
        row = 1
        for src in img_src:
            urllib.request.urlretrieve(src, file_path + '/img/Instagram_' + str(row) + ".jpeg")
            row += 1
            if (row % 10 == 0):
                print("(" + str(row) + "/" + str(len(img_src)) + ") Images Downloaded")

        print("Closing Instagram")
        driver.quit()






def Create_Dir2(dir_name):
        if not os.path.exists("data1"):
            try:
                os.mkdir("data1")
                print("Created directory 'data1'")
            except:
                print("Unable to create directory 'data1': Directory already exists")
        else:
            print("Unable to create directory 'data1': Directory already exists")

        if not os.path.exists("data1/data1_" + dir_name):
            try:
                os.mkdir("data1/data1_" + dir_name)
                print("Created directory 'data1/data1_" + dir_name + "'")
            except:
                print("Unable to create directory 'data1/data1_" + dir_name + "': Directory already exists")
        else:
            print("Unable to create directory 'data1/data1_" + dir_name + "': Directory already exists")

def Scrape_Twitter(Consumer_Key, Consumer_Secret, Access_Token, Access_Token_Secret, tag, limit, lang='en'):
        Create_Dir2(tag)

        print("Starting Scrapping Twitter")
        file_path = "data1/data1_" + tag

        auth = tweepy.OAuthHandler(Consumer_Key, Consumer_Secret)
        auth.set_access_token(Access_Token, Access_Token_Secret)
        api = tweepy.API(auth)

        tweets = tweepy.Cursor(api.search, q=tag, lang=lang).items(limit)

        # Create a workbook for excel
        tag_File = file_path + "/" + tag + "_Twitter.xlsx"
        wb = openpyxl.Workbook()
        ws_Captions = wb.create_sheet(title="Posts")
        col = 'A'
        row = 1

        hashtags = {}
        ext_links = []
        for tweet in tweets:
            text = tweet.text.lower()
            ws_Captions[col + str(row)] = text

            for tag in text.split():
                if tag.startswith("#"):
                    if tag[1:] not in hashtags:
                        hashtags[tag[1:]] = 1
                    elif tag[1:] in hashtags:
                        hashtags[tag[1:]] = hashtags[tag[1:]] + 1
                if tag[:4] == 'http':
                    ext_links.append(tag)



            if (row % 50 == 0):
                print("Dumped " + str(row) + " Tweets")
            row += 1


        hashtags = sorted(hashtags.items(), reverse=True)

        ws_Tags = wb.create_sheet(title="Tags")
        tagName = 'A'
        tagFreq = 'B'
        row = 1

        print("Dumping Related Hashtags")
        for tag in hashtags:
            ws_Tags[tagName + str(row)] = tag[0]
            ws_Tags[tagFreq + str(row)] = tag[1]
            row += 1


        print("Dumping External Links")
        ws_Links = wb.create_sheet(title="Links")
        row = 1
        for link in ext_links:
            ws_Links['A' + str(row)] = link
            row += 1


        wb.save(tag_File)

        time.sleep(5)



        print("Closing Twitter")
#-----------------------------------------------------------------------------------------------


#
# #
# # FOR NEWS RIVER
# print("----------------:INPUT FOR NEWS RIVER WEBSITE SCRAPER:--------------------- ")
# keyword = int(input("Do you want to search for more than one Keyword:\n1 for Yes\n2 for No"))
# limit = int(input("How Many post you want to fetch:"))
# fil = int(input("Do you want to Filter Date:\n2 For  NO:\n1 For YES:"))
# if keyword == 1 and fil == 1:
#     date2 = input("From Date (YYYY-MM-DD):")
#     date3 = input("To Date (YYYY-MM-DD):")
#     text1 = input("Enter Keywords to Search use AND/OR between keywords:")
#     txt1 = text1.split()
#     oper = txt1[1].upper()
#     jsonf=key1fil1(txt1,oper,date2,date3,limit)
#
# elif keyword == 1 and fil == 2:
#     text1 = input("Enter Keywords to Search use AND/OR between keywords:")
#     txt1 = text1.split()
#     oper = txt1[1].upper()
#     jsonf=key1fill2(txt1,oper,limit)
#
# elif keyword == 2 and fil == 1:
#     date2 = input("From Date (YYYY-MM-DD):")
#     date3 = input("To Date (YYYY-MM-DD):")
#     text1 = input("Enter Keyword to search")
#     jsonf=key2fill1(text1,date2,date3,limit)
#
# elif keyword == 2 and fil == 2:
#     text1 = input("Enter Keyword to search")
#     jsonf=key2fill2(text1,limit)
#
# else:
#     print("Wrong Input")
#
# # FOR GOOGLE
# print("----------------:INPUT FOR GOOGLE NEWS SCRAPER:--------------------- ")
# keygi = input("Enter Keyword")
# dateg = int(input("Do you want to filter date:\n1 For YES\n2 For NO"))
# if dateg == 1:
#     a = int(input("Please Select Filter:\n1)Past Hour:Press 1\n2)Past 24 Hour:Press 2\n3)Past Week:Press 3\n4)Past Year:Press 4 "))
#     if a == 1:
#         responseg2 = keygdateg1(keygi)
#
#     elif a == 2:
#         responseg2 = keygdateg2(keygi)
#     elif a == 3:
#         responseg2 = keygdateg3(keygi)
#     elif a == 4:
#         responseg2 = keygdateg4(keygi)
#     else:
#         print("Worng Input")
# elif dateg == 2:
#     responseg2 = keyg2r(keygi)
# # else:
# #     print("Wrong Input")
#
# # soup = BeautifulSoup(responseg2, 'lxml')
# soup2 = BeautifulSoup(responseg2,'lxml')
#
# print("------------------------------Instagam-------------------------------------")
# print("Starting Social Media Scrapper...")
# keywordin = str(input("Enter keyword to search for images on Instagram: "))
# insta_limit = int(input("Enter how many posts to scrape from Instagram: "))
#
#
# print("-------------------------------   Twitter Scraper    --------------------------------------")
# consumerKey = '2TdOT1UVvFPc4EbeDOBMYJffK'
# consumerSecret = 'vrF3KZdgjPdczFaAQ7iYegR1Kd3YCVCKxnzYAJ1Yy47SczeRbl'
# accessToken = '1143437697920335874-Kg63Sf9sozQP82itGo9IWYBIfulL7j'
# accessTokenSecret = 'b0Sz7uJVyQuIxb6GllwjgiY2wHMA8FA0ZKsUW9Pa76ZPr'
#
# keywordtw = str(input("Enter keyword to search for: "))
# twitter_limit = int(input("Enter how many posts to scrape from Twitter: "))


# soup2 = BeautifulSoup(responseg2,'lxml')
#
# t1 = threading.Thread(target=forg, args=(soup2,))
# t2 = threading.Thread(target=forn, args=(jsonf,))
# t3 = threading.Thread(target=Scrape_Instagram,args=(keywordin,insta_limit,'chrome',))

# t4 = threading.Thread(target=Scrape_Twitter,args=(consumerKey,consumerSecret,accessToken,accessTokenSecret,keywordtw,twitter_limit,))

# t1.start()
# t2.start()
# t3.start()
# t4.start()